# -*- coding: utf-8 -*-
"""
Carga el catálogo del bot: lista de precios del proveedor (Excel) cruzada contra
el inventario real de Heliteb (Odoo).

    python scripts/cargar_catalogo.py --dsn "postgresql://usuario:clave@host:puerto/heliteb"

Fuentes:
  - Excel  "PRUEBA TECNICA AGENTE AI.xlsx", hojas LISTA PRECIOS HIKVISION / EZVIZ.
    Es lo que el PROVEEDOR vende: una fila por variante (lente), con código SAP y MSRP.
  - Odoo   product.template + stock.quant vía XML-RPC. Es lo que NOSOTROS tenemos:
    una ficha por familia (el SKU no distingue el lente) y existencias por bodega.

El cruce se hace en tres niveles, de más a menos fiable: código SAP de Odoo ->
SKU idéntico al modelo -> SKU igual al modelo ignorando el lente. El último puede
ser 1 SKU para varias variantes del Excel; esos quedan marcados exclusivo=FALSE
para que el agente no afirme que tiene justo esa variante (ver 003_stock_odoo.sql).

Por defecto NO escribe: imprime el resumen y sale. Usar --escribir para aplicar.
"""
import argparse
import json
import os
import re
import sys
from collections import defaultdict

import openpyxl
import psycopg2
import psycopg2.extras

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_POR_DEFECTO = os.path.join(RAIZ, "InformacionEmpresa", "PRUEBA TECNICA AGENTE AI.xlsx")

# Bodegas de mostrador: lo que un cliente puede recoger hoy. Son las 10 sucursales
# de venta, las mismas que siembra sql/schema.sql en la tabla `bodegas`.
BODEGAS_SEDE = {"01", "03", "04", "08", "09", "11", "17", "21", "25", "26"}
# Bodegas centrales: hay existencias, pero requieren traslado a la sede.
BODEGAS_CENTRALES = {"LS", "CDBAQ", "CDBOG", "DLTD", "Mini", "07", "00"}


# ---------------------------------------------------------------------------
# Normalización
# ---------------------------------------------------------------------------
def norm_sap(valor):
    """Código SAP a solo dígitos. Descarta la basura (en Odoo y en el panel de
    abastecimiento hay filas donde copiaron el SKU dentro del campo SAP)."""
    if not valor:
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    solo_digitos = re.sub(r"\D", "", texto)
    return solo_digitos if len(solo_digitos) >= 6 else ""


def norm_full(valor):
    """Modelo completo en alfanumérico: distingue lente y región."""
    if not valor:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(valor).upper())


def norm_base(valor):
    """Modelo sin lente ni región: DS-2CE16D0T-IRF(2.8mm)(C) -> DS2CE16D0TIRF."""
    if not valor:
        return ""
    texto = str(valor).upper().split("|")[0]
    texto = re.sub(r"\([^)]*\)", " ", texto)
    texto = re.sub(r"\b\d+(\.\d+)?\s*-?\s*(\d+(\.\d+)?)?\s*MM\b", " ", texto)
    return re.sub(r"[^A-Z0-9]", "", texto)


def texto(valor):
    if valor is None:
        return None
    limpio = str(valor).strip()
    return limpio or None


def numero(valor):
    if valor is None or str(valor).strip() == "":
        return None
    try:
        return round(float(valor), 2)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Lectura del Excel del proveedor
# ---------------------------------------------------------------------------
def leer_excel(ruta):
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    filas = []

    for fila in libro["LISTA PRECIOS HIKVISION"].iter_rows(min_row=2, values_only=True):
        fila = list(fila) + [None] * 14
        sap, _marca, categoria, linea, serie, subserie, modelo = fila[0:7]
        if not sap and not modelo:
            continue
        filas.append(dict(
            sap=norm_sap(sap), marca="HIKVISION", categoria=texto(categoria),
            linea=texto(linea), serie=texto(serie), sub_serie=texto(subserie),
            modelo=texto(modelo), parametro_1=texto(fila[8]), parametro_2=texto(fila[9]),
            parametro_3=texto(fila[10]), descripcion=texto(fila[11]),
            etiqueta=texto(fila[12]), msrp=numero(fila[13])))

    categoria_actual = None
    for fila in libro["LISTA PRECIOS EZVIZ"].iter_rows(min_row=6, values_only=True):
        fila = list(fila) + [None] * 5
        imagen, referencia, descripcion, sap, msrp = fila[0:5]
        # Las filas de sección traen solo la primera celda ("CAMARAS WIFI INTERIORES").
        if imagen and all(v is None or str(v).strip() == "" for v in (referencia, descripcion, sap, msrp)):
            categoria_actual = texto(imagen)
            continue
        if not referencia and not sap:
            continue
        filas.append(dict(
            sap=norm_sap(sap), marca="EZVIZ", categoria=categoria_actual, linea=None,
            serie=None, sub_serie=None, modelo=texto(referencia), parametro_1=None,
            parametro_2=None, parametro_3=None, descripcion=texto(descripcion),
            etiqueta=None, msrp=numero(msrp)))
    libro.close()

    # Un mismo SAP repetido en el Excel: nos quedamos con la primera aparición.
    vistos, unicos, descartados = set(), [], 0
    for fila in filas:
        if not fila["sap"]:
            descartados += 1
            continue
        if fila["sap"] in vistos:
            descartados += 1
            continue
        vistos.add(fila["sap"])
        unicos.append(fila)
    return unicos, descartados


# ---------------------------------------------------------------------------
# Lectura de Odoo (en vivo o desde un snapshot ya extraído)
# ---------------------------------------------------------------------------
def leer_odoo(snapshot=None):
    if snapshot:
        with open(snapshot, encoding="utf-8") as fh:
            return json.load(fh)
    return extraer_de_odoo()


def extraer_de_odoo():
    import xmlrpc.client

    cfg = {}
    ruta_env = os.environ.get("ODOO_ENV", r"c:\Users\JoseArmando\Documents\odoo-mcp\mcp-server\.env")
    with open(ruta_env, encoding="utf-8") as fh:
        for linea in fh:
            linea = linea.strip()
            if linea and not linea.startswith("#") and "=" in linea:
                clave, valor = linea.split("=", 1)
                cfg[clave.strip()] = valor.strip()

    url, base = cfg["ODOO_URL"], cfg["ODOO_DB"]
    usuario, clave = cfg["ODOO_USER"], cfg["ODOO_PASSWORD"]
    comun = xmlrpc.client.ServerProxy("%s/xmlrpc/2/common" % url, allow_none=True)
    uid = comun.authenticate(base, usuario, clave, {})
    modelos = xmlrpc.client.ServerProxy("%s/xmlrpc/2/object" % url, allow_none=True)

    def eje(modelo, metodo, args=None, **kw):
        return modelos.execute_kw(base, uid, clave, modelo, metodo, args or [], kw)

    def paginar(modelo, dominio, campos):
        salida, desplazamiento = [], 0
        while True:
            lote = eje(modelo, "search_read", [dominio], fields=campos,
                       limit=500, offset=desplazamiento, order="id asc")
            if not lote:
                return salida
            salida.extend(lote)
            desplazamiento += 500

    ubicaciones = {}
    for loc in eje("stock.location", "search_read", [[["usage", "=", "internal"]]],
                   fields=["id", "complete_name", "warehouse_id"], limit=500):
        nombre = loc["complete_name"] or ""
        ubicaciones[str(loc["id"])] = dict(
            codigo=nombre.split("/")[0].strip() if "/" in nombre else nombre.strip(),
            nombre=nombre,
            warehouse=loc["warehouse_id"][1] if loc.get("warehouse_id") else None)

    plantillas = paginar("product.template", [["active", "in", [True, False]]],
                         ["id", "default_code", "name", "x_studio_cdigo_sap", "brand_name",
                          "standard_price", "active"])
    variantes = paginar("product.product", [["active", "in", [True, False]]],
                        ["id", "product_tmpl_id"])
    var2tmpl = {v["id"]: (v["product_tmpl_id"][0] if v.get("product_tmpl_id") else None)
                for v in variantes}

    acumulado, desplazamiento = defaultdict(float), 0
    while True:
        grupo = eje("stock.quant", "read_group",
                    [[["location_id.usage", "=", "internal"]], ["quantity:sum"],
                     ["product_id", "location_id"]],
                    lazy=False, limit=2000, offset=desplazamiento)
        if not grupo:
            break
        for reg in grupo:
            pid = reg["product_id"][0] if reg.get("product_id") else None
            lid = reg["location_id"][0] if reg.get("location_id") else None
            tmpl = var2tmpl.get(pid)
            if tmpl and str(lid) in ubicaciones:
                acumulado[(tmpl, lid)] += float(reg.get("quantity") or 0)
        desplazamiento += 2000

    return dict(ubicaciones=ubicaciones, plantillas=plantillas,
                existencias=[{"tmpl": t, "loc": l, "cant": c}
                             for (t, l), c in acumulado.items() if c != 0])


def tipo_de_bodega(codigo, nombre):
    mayus = (nombre or "").upper()
    if codigo in BODEGAS_SEDE:
        return "sede"
    if "GARANT" in mayus:
        return "garantias"
    if codigo in BODEGAS_CENTRALES or "LOGISTICA" in mayus or "CEDIS" in mayus:
        return "central"
    return "otra"


# ---------------------------------------------------------------------------
# Armado del cruce
# ---------------------------------------------------------------------------
def construir(proveedor, odoo, alcance):
    ubicaciones = odoo["ubicaciones"]
    plantillas = odoo["plantillas"]

    existencias_por_tmpl = defaultdict(list)
    for reg in odoo["existencias"]:
        existencias_por_tmpl[reg["tmpl"]].append(reg)

    # Un mismo default_code puede tener varias fichas en Odoo (producto recreado).
    # Se consolidan en un solo SKU sumando existencias, y se marca duplicado_odoo.
    fichas_por_sku = defaultdict(list)
    for plantilla in plantillas:
        codigo = (plantilla.get("default_code") or "").strip()
        if codigo:
            fichas_por_sku[codigo].append(plantilla)

    items = {}
    for sku, fichas in fichas_por_sku.items():
        principal = max(fichas, key=lambda p: (bool(p.get("active")), p["id"]))
        sap_odoo = ""
        for ficha in fichas:
            sap_odoo = sap_odoo or norm_sap(ficha.get("x_studio_cdigo_sap"))
        por_bodega = defaultdict(float)
        for ficha in fichas:
            for reg in existencias_por_tmpl.get(ficha["id"], []):
                info = ubicaciones[str(reg["loc"])]
                por_bodega[info["nombre"]] += reg["cant"]
        items[sku] = dict(
            sku=sku, odoo_tmpl_id=principal["id"], nombre=texto(principal.get("name")),
            marca=texto(principal.get("brand_name")), codigo_sap_odoo=sap_odoo or None,
            costo=numero(principal.get("standard_price")),
            activo=any(bool(f.get("active")) for f in fichas),
            duplicado_odoo=len(fichas) > 1,
            existencias=[dict(nombre=nombre, cantidad=round(cant, 2))
                         for nombre, cant in por_bodega.items() if cant != 0])

    info_por_nombre = {v["nombre"]: v for v in ubicaciones.values()}

    por_sap, por_full, por_base = defaultdict(list), defaultdict(list), defaultdict(list)
    for item in items.values():
        if item["codigo_sap_odoo"]:
            por_sap[item["codigo_sap_odoo"]].append(item)
        por_full[norm_full(item["sku"])].append(item)
        por_base[norm_base(item["sku"])].append(item)

    # Primera pasada: a qué SKU apunta cada fila del proveedor.
    for fila in proveedor:
        if fila["sap"] and fila["sap"] in por_sap:
            fila["_skus"], fila["_metodo"] = por_sap[fila["sap"]], "sap"
        elif norm_full(fila["modelo"]) in por_full:
            fila["_skus"], fila["_metodo"] = por_full[norm_full(fila["modelo"])], "modelo"
        elif norm_base(fila["modelo"]) in por_base:
            fila["_skus"], fila["_metodo"] = por_base[norm_base(fila["modelo"])], "modelo_base"
        else:
            fila["_skus"], fila["_metodo"] = [], None

    # Segunda pasada: un SKU reclamado por varias filas del Excel NO es exclusivo de
    # ninguna de ellas, ni siquiera de la que cruzó por SAP. Aunque Odoo le asigne el
    # SAP del lente de 2.8mm, las unidades físicas son UNA sola pila que respalda a
    # las dos variantes: marcar como exclusiva la del SAP haría que el agente reporte
    # 25 unidades de 2.8mm y otras 25 de 3.6mm, cuando en bodega hay 25 en total.
    reclamos = defaultdict(list)
    for fila in proveedor:
        for item in fila["_skus"]:
            reclamos[item["sku"]].append(fila)
    for fila in proveedor:
        fila["_exclusivo"] = all(len(reclamos[item["sku"]]) == 1 for item in fila["_skus"])

    # Productos que tenemos en bodega y el proveedor no lista.
    cruzados = {item["sku"] for fila in proveedor for item in fila["_skus"]}
    huerfanos = []
    for item in items.values():
        if item["sku"] in cruzados:
            continue
        vendible = sum(e["cantidad"] for e in item["existencias"]
                       if tipo_de_bodega(info_por_nombre[e["nombre"]]["codigo"],
                                         e["nombre"]) in ("sede", "central"))
        if vendible <= 0:
            continue
        marca = (item["marca"] or "").upper()
        if alcance == "b":
            continue
        if alcance == "a" and marca not in ("HIKVISION", "EZVIZ", "HILOOK"):
            continue
        huerfanos.append(item)

    return items, huerfanos, info_por_nombre


# ---------------------------------------------------------------------------
# Escritura en Postgres
# ---------------------------------------------------------------------------
def escribir(dsn, proveedor, items, huerfanos, info_por_nombre):
    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    cur = conn.cursor()

    migracion = os.path.join(RAIZ, "sql", "migrations", "003_stock_odoo.sql")
    with open(migracion, encoding="utf-8") as fh:
        cur.execute(fh.read())

    # --- marcas y categorías ---
    marcas = {f["marca"] for f in proveedor} | {(i["marca"] or "SIN MARCA") for i in huerfanos}
    for marca in sorted(marcas):
        cur.execute("INSERT INTO marcas (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING",
                    (marca[:50],))
    cur.execute("SELECT nombre, id_marca FROM marcas")
    id_marca = dict(cur.fetchall())

    categorias = {f["categoria"] for f in proveedor if f["categoria"]}
    for categoria in sorted(categorias):
        cur.execute("""INSERT INTO categorias (nombre, id_padre) VALUES (%s, NULL)
                       ON CONFLICT DO NOTHING""", (categoria[:100],))
    cur.execute("SELECT nombre, id_categoria FROM categorias WHERE id_padre IS NULL")
    id_categoria = dict(cur.fetchall())

    # --- productos del proveedor ---
    psycopg2.extras.execute_batch(cur, """
        INSERT INTO productos (codigo_sap, id_marca, id_categoria, linea, serie, sub_serie,
                               modelo, parametro_1, parametro_2, parametro_3, descripcion,
                               modelo_etiqueta, origen, activo)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'proveedor',TRUE)
        ON CONFLICT (codigo_sap) DO UPDATE SET
            id_marca = EXCLUDED.id_marca, id_categoria = EXCLUDED.id_categoria,
            linea = EXCLUDED.linea, serie = EXCLUDED.serie, sub_serie = EXCLUDED.sub_serie,
            modelo = EXCLUDED.modelo, parametro_1 = EXCLUDED.parametro_1,
            parametro_2 = EXCLUDED.parametro_2, parametro_3 = EXCLUDED.parametro_3,
            descripcion = EXCLUDED.descripcion, modelo_etiqueta = EXCLUDED.modelo_etiqueta,
            origen = 'proveedor', fecha_actualizacion = now()
    """, [(f["sap"], id_marca[f["marca"]], id_categoria.get(f["categoria"]), f["linea"],
           f["serie"], f["sub_serie"], f["modelo"] or f["sap"], f["parametro_1"],
           f["parametro_2"], f["parametro_3"], f["descripcion"], f["etiqueta"])
          for f in proveedor], page_size=500)

    psycopg2.extras.execute_batch(cur, """
        INSERT INTO precios (codigo_sap, precio_msrp_cop) VALUES (%s, %s)
        ON CONFLICT (codigo_sap) DO UPDATE SET
            precio_msrp_cop = EXCLUDED.precio_msrp_cop, fecha_actualizacion = now()
    """, [(f["sap"], f["msrp"]) for f in proveedor if f["msrp"] is not None], page_size=500)

    # --- productos que solo tenemos nosotros (no están en la lista del proveedor) ---
    if huerfanos:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO productos (codigo_sap, id_marca, modelo, descripcion, origen, activo)
            VALUES (%s,%s,%s,%s,'stock_local',TRUE)
            ON CONFLICT (codigo_sap) DO UPDATE SET
                modelo = EXCLUDED.modelo, descripcion = EXCLUDED.descripcion,
                fecha_actualizacion = now()
        """, [(i["sku"][:50], id_marca[i["marca"] or "SIN MARCA"], i["sku"][:100],
               i["nombre"]) for i in huerfanos], page_size=500)

    # --- inventario real ---
    cur.execute("TRUNCATE stock_existencias, producto_stock")
    cur.execute("DELETE FROM stock_items")

    usados = {i["sku"] for f in proveedor for i in f["_skus"]} | {i["sku"] for i in huerfanos}
    psycopg2.extras.execute_batch(cur, """
        INSERT INTO stock_items (sku, odoo_tmpl_id, nombre, marca, codigo_sap_odoo, costo,
                                 activo, duplicado_odoo)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, [(i["sku"][:80], i["odoo_tmpl_id"], i["nombre"], (i["marca"] or "")[:80] or None,
           i["codigo_sap_odoo"], i["costo"], i["activo"], i["duplicado_odoo"])
          for k, i in items.items() if k in usados], page_size=500)

    filas_exist = []
    for sku in usados:
        item = items[sku]
        for existencia in item["existencias"]:
            info = info_por_nombre[existencia["nombre"]]
            filas_exist.append((sku[:80], info["codigo"][:20], existencia["nombre"][:120],
                                tipo_de_bodega(info["codigo"], existencia["nombre"]),
                                existencia["cantidad"]))
    psycopg2.extras.execute_batch(cur, """
        INSERT INTO stock_existencias (sku, codigo_bodega, nombre_bodega, tipo_bodega, cantidad)
        VALUES (%s,%s,%s,%s,%s) ON CONFLICT (sku, nombre_bodega) DO NOTHING
    """, filas_exist, page_size=500)

    puente = []
    for fila in proveedor:
        for item in fila["_skus"]:
            puente.append((fila["sap"], item["sku"][:80], fila["_metodo"], fila["_exclusivo"]))
    for item in huerfanos:
        puente.append((item["sku"][:50], item["sku"][:80], "propio", True))
    psycopg2.extras.execute_batch(cur, """
        INSERT INTO producto_stock (codigo_sap, sku, metodo, exclusivo)
        VALUES (%s,%s,%s,%s) ON CONFLICT (codigo_sap, sku) DO NOTHING
    """, puente, page_size=500)

    # --- inventario (tabla vieja): solo sedes y cruces exclusivos, para que las
    #     vistas que ya existían (vista_stock) no reporten inventario inflado ---
    cur.execute("DELETE FROM inventario")
    cur.execute("SELECT codigo_bodega, id_bodega FROM bodegas")
    id_bodega = dict(cur.fetchall())
    filas_inv = defaultdict(float)
    for fila in proveedor:
        if not fila["_exclusivo"]:
            continue
        for item in fila["_skus"]:
            for existencia in item["existencias"]:
                info = info_por_nombre[existencia["nombre"]]
                if info["codigo"] in id_bodega and tipo_de_bodega(info["codigo"], existencia["nombre"]) == "sede":
                    filas_inv[(fila["sap"], id_bodega[info["codigo"]])] += existencia["cantidad"]
    psycopg2.extras.execute_batch(cur, """
        INSERT INTO inventario (codigo_sap, id_bodega, cantidad_disponible) VALUES (%s,%s,%s)
        ON CONFLICT (codigo_sap, id_bodega) DO UPDATE SET
            cantidad_disponible = EXCLUDED.cantidad_disponible, fecha_actualizacion = now()
    """, [(sap, bod, int(cant)) for (sap, bod), cant in filas_inv.items() if cant > 0],
        page_size=500)

    conn.commit()
    cur.close()
    conn.close()
    return len(filas_exist), len(puente), len(filas_inv)


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dsn", default=os.environ.get("HELITEB_DSN"),
                        help="postgresql://usuario:clave@host:puerto/heliteb")
    parser.add_argument("--xlsx", default=XLSX_POR_DEFECTO)
    parser.add_argument("--snapshot", help="JSON ya extraído de Odoo (evita reconsultar)")
    # Por defecto solo el Excel: HIKVISION y EZVIZ son las unicas marcas con ficha
    # tecnica completa (categoria, linea, serie, parametros, descripcion), y sin eso
    # el agente no puede recomendar - solo repetir un nombre. Las otras marcas del
    # inventario (POWEST, TP-LINK, RUIJIE...) traen nombre y costo, nada mas.
    parser.add_argument("--alcance", choices=["a", "b", "c"], default="b",
                        help="b: solo el Excel del proveedor, HIKVISION+EZVIZ (por defecto) | "
                             "a: ademas el stock propio HIK/EZVIZ/HILOOK sin ficha | "
                             "c: ademas todo el stock de las demas marcas")
    parser.add_argument("--escribir", action="store_true", help="aplica los cambios")
    args = parser.parse_args()

    proveedor, descartadas = leer_excel(args.xlsx)
    print("Excel del proveedor: %d productos con SAP (%d filas descartadas)" % (
        len(proveedor), descartadas))

    odoo = leer_odoo(args.snapshot)
    print("Odoo: %d fichas, %d existencias, %d ubicaciones internas" % (
        len(odoo["plantillas"]), len(odoo["existencias"]), len(odoo["ubicaciones"])))

    items, huerfanos, info_por_nombre = construir(proveedor, odoo, args.alcance)

    metodos = defaultdict(int)
    estados = defaultdict(int)
    no_exclusivos = 0
    for fila in proveedor:
        metodos[fila["_metodo"] or "sin_cruce"] += 1
        if not fila["_exclusivo"]:
            no_exclusivos += 1
        sede = central = 0.0
        for item in fila["_skus"]:
            for existencia in item["existencias"]:
                info = info_por_nombre[existencia["nombre"]]
                tipo = tipo_de_bodega(info["codigo"], existencia["nombre"])
                if tipo == "sede":
                    sede += existencia["cantidad"]
                elif tipo == "central":
                    central += existencia["cantidad"]
        if sede > 0:
            estados["EN_SEDE"] += 1
        elif central > 0:
            estados["EN_BODEGA_CENTRAL"] += 1
        elif fila["_skus"]:
            estados["AGOTADO"] += 1
        else:
            estados["BAJO_PEDIDO"] += 1

    print("\nCRUCE:")
    for clave in ("sap", "modelo", "modelo_base", "sin_cruce"):
        print("   %-12s %5d" % (clave, metodos[clave]))
    print("   variantes NO exclusivas (stock de familia): %d" % no_exclusivos)
    print("\nDISPONIBILIDAD:")
    for clave, valor in sorted(estados.items(), key=lambda x: -x[1]):
        print("   %-20s %5d  (%.1f%%)" % (clave, valor, 100.0 * valor / len(proveedor)))
    print("\nProductos propios fuera de la lista del proveedor: %d" % len(huerfanos))
    duplicados = [i for i in items.values() if i["duplicado_odoo"]]
    print("Fichas duplicadas en Odoo detectadas: %d" % len(duplicados))

    if not args.escribir:
        print("\n(simulación: no se escribió nada. Usar --escribir para aplicar)")
        return 0
    if not args.dsn:
        print("\nFalta --dsn (o la variable HELITEB_DSN)", file=sys.stderr)
        return 1

    n_exist, n_puente, n_inv = escribir(args.dsn, proveedor, items, huerfanos, info_por_nombre)
    print("\nESCRITO: %d productos, %d existencias, %d cruces, %d filas de inventario" % (
        len(proveedor) + len(huerfanos), n_exist, n_puente, n_inv))
    return 0


if __name__ == "__main__":
    sys.exit(main())
